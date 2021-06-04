#!/usr/bin/env python
# coding: utf-8

import pandas as pd
import openpyxl as opx

def write_to_existed_sheet(df, file_path, sheet_name, start_row=1, start_col=1, use_col=True, use_index=True):
    """
    The function to write to an existed excel sheet of your excel file. 

    df:pandas DataFrame
    file_path: str
        the name of your excel file (pleasu use ".xlsx" not ".xls")
    sheet_name: str
        the sheet name you want to write to
    start_row: int
        the row number of top left cell of the table 
    start_col: int
        the column number of top left cell of the table
    use_col: Boolean
        whether to write out the column name
    use_index:Boolean
        whether to write out the index name
    """
    
    wb = opx.load_workbook(file_path)
    ws = wb[sheet_name]

    if use_col:
        for i, col_name in enumerate(df.columns):
            ws.cell(row=start_row, column=start_col+1+i, value=col_name)
        start_row_adj = start_row+1

    if use_index:
        for j, row_name in enumerate(df.index):
            ws.cell(row=start_row+1+j, column=start_col, value=row_name)
        start_col += 1

    for row_num in range(len(df)):
        for col_num, value in enumerate(df.iloc[row_num].to_list()):
            ws.cell(row=start_row_adj+row_num,
                    column=start_col+col_num, value=value)
    wb.save(file_path)

